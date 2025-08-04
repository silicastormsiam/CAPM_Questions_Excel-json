# File Name: convert_questions.py
# Owner: Andrew
# Purpose: Converts CAPM questions from test_questions.xlsx to questions.json, appending non-destructively
# Version Control: v1.0
# Change Log: Created: 04-Aug-2025, v1.0: Initial script for 3 questions

import openpyxl
import json

def convert_questions(excel_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\test_questions.xlsx', 
                     json_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\data\\questions.json'):
    try:
        with open(json_file, 'r') as f:
            json_data = json.load(f)
        last_id = max(item['id'] for item in json_data) if json_data else 301
    except FileNotFoundError:
        json_data = []
        last_id = 301

    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet1']

    new_questions = []
    for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row, values_only=True):
        new_questions.append({
            "id": last_id + 1,
            "type": row[1],
            "domain": row[2],
            "module": str(row[3]),
            "question": row[4],
            "options": [f"a. {row[5]}", f"b. {row[6]}", f"c. {row[7]}", f"d. {row[8]}"],
            "answer": [int(x) for x in row[9].split(',')] if ',' in str(row[9]) else int(row[9]) if row[9] else 0,
            "explanation": row[10]
        })
        last_id += 1

    json_data.extend(new_questions)
    with open(json_file, 'w') as f:
        json.dump(json_data, f, indent=4)
    print(f"Added {len(new_questions)} questions to {json_file}")

if __name__ == "__main__":
    convert_questions()