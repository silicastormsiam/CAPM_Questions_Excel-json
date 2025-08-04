# File Name: add_questions_to_excel.py
# Owner: Andrew
# Purpose: Appends CAPM questions to test_questions.xlsx, scalable for any number
# Version Control: v1.0
# Change Log: Created: 04-Aug-2025, v1.0: Initial script for 3 questions

import openpyxl

def append_questions(questions, excel_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\test_questions.xlsx'):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet1']
    
    # Start at last non-empty row or row 5
    start_row = max(ws.max_row + 1, 5)
    
    for i, q in enumerate(questions, start=start_row):
        ws[f'A{i}'] = q['id']
        ws[f'B{i}'] = q['type']
        ws[f'C{i}'] = q['domain']
        ws[f'D{i}'] = q['module']
        ws[f'E{i}'] = q['question']
        ws[f'F{i}'] = q['option_a']
        ws[f'G{i}'] = q['option_b']
        ws[f'H{i}'] = q['option_c']
        ws[f'I{i}'] = q['option_d']
        ws[f'J{i}'] = q['answer']
        ws[f'K{i}'] = q['explanation']
    
    wb.save(excel_file)
    print(f"Appended {len(questions)} CAPM questions")

if __name__ == "__main__":
    questions = [
        {
            "id": 302,
            "type": "single",
            "domain": 4,
            "module": "6",
            "question": "Who is the project sponsor?",
            "option_a": "The person involved in the project who has a vested interest in the project's success",
            "option_b": "The person who will use the product produced by a project",
            "option_c": "The person who plans, organizes, and oversees the project",
            "option_d": "The person who is accountable for the project and ensures the project delivers the agreed-upon value to a business",
            "answer": 3,  # D = 3
            "explanation": "A project sponsor is the person who is accountable for the project. They ensure the project delivers the agreed-upon value to the business. ECO Domain 4, Task 4.2; PMBOK 7th Edition, Section 4.2."
        },
        {
            "id": 303,
            "type": "multiple",
            "domain": 4,
            "module": "6",
            "question": "When choosing a project team, a project manager considers required roles, team size, and which three additional factors?",
            "option_a": "Availability",
            "option_b": "Motivation",
            "option_c": "Necessary skills",
            "option_d": "Stakeholder preferences",
            "answer": "0,1,2",  # A,B,C
            "explanation": "Everyone on the team needs to be available to join the project. This means they are not staffed on another big project and have the time to contribute. Beyond availability, team members also need to have key skills and be motivated to complete their tasks. Everyone on the team needs to have the right skills to do the job. If they don’t have the necessary skills, the project manager should ensure that they receive timely training to avoid project delays. Beyond key skills, team members also need to be available and motivated. Stakeholder preferences about team members are not a factor project managers need to consider when building the team. ECO Domain 4, Task 4.2; PMBOK 7th Edition, Section 4.2."
        },
        {
            "id": 304,
            "type": "multiple",
            "domain": 4,
            "module": "6",
            "question": "Which two factors are determined on a stakeholder power grid? Select all that apply.",
            "option_a": "Influence",
            "option_b": "Interest",
            "option_c": "Input",
            "option_d": "Impact",
            "answer": "0,1",  # A,B
            "explanation": "The stakeholder power grid is a two-by-two grid used for conducting a stakeholder analysis. The power grid assigns each stakeholder a level of importance to the project using two measures: interest and influence. ECO Domain 4, Task 4.2; PMBOK 7th Edition, Section 4.2."
        }
    ]
    append_questions(questions)